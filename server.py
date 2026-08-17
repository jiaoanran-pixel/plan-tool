#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""车辆计划统计与对账系统 - 本地服务

运行：python3 server.py
手机/电脑浏览器访问：http://<本机IP>:8766
"""

import base64
import datetime
import json
import mimetypes
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import uuid
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qs, urlparse

ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(ROOT, "data"))
IMAGES_DIR = os.path.join(DATA_DIR, "images")
DB_PATH = os.path.join(DATA_DIR, "plans.db")
STATIC_DIR = os.path.join(ROOT, "static")
OCR_BIN = os.path.join(ROOT, "ocr_helper")
OCR_SRC = os.path.join(ROOT, "ocr_helper.m")

PORT = int(os.environ.get("PORT", "8766"))
HOST = os.environ.get("HOST", "0.0.0.0")

DOC_TYPES = {"load": "装车磅单", "unload": "卸车磅单", "waybill": "运单"}
IMG_COLUMNS = {"load": "img_load", "unload": "img_unload", "waybill": "img_waybill"}

# ---------------- 正则 ----------------
PLATE_RE = re.compile(
    r"[京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤青藏川宁琼]"
    r"[A-Z][A-Z0-9]{4,6}挂?"
)
FULL_DATE_RE = re.compile(r"(20\d{2})[-/年.](\d{1,2})[-/月.](\d{1,2})")
CN_DAY_RE = re.compile(r"(?<![\d])(\d{1,2})[号日]")
PRICE_RE = re.compile(r"价格[：:\s]*([0-9]+(?:\.[0-9]+)?)")
SUPPLIER_RE = re.compile(r"供应商[：:\s]*([^\s；;，,]+)")
NET_WEIGHT_RE = re.compile(
    r"净重[：:\s]*([0-9]+(?:\.[0-9]+)?)\s*(吨|t|T|kg|KG|千克|kg)?"
)
ID_RE = re.compile(r"\d{17}[\dXx]")
PHONE_RE = re.compile(r"1[3-9]\d{9}")
NAME_RE = re.compile(r"^[\u4e00-\u9fff]{2,4}$")
COMPANY_RE = re.compile(r"([\u4e00-\u9fff]{2,}(?:运输|物流|公司)[\u4e00-\u9fff]{0,12})")
# 常用站点词典：后续新增站点时只需在此处补充。
COMMON_STATIONS = ("宜章西西站", "宜章西东站", "鲁塘坳", "湘阴渡", "汝城南")


# ---------------- 数据库 ----------------
def get_db():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(IMAGES_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    os.makedirs(IMAGES_DIR, exist_ok=True)
    with get_db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS plans (
                id TEXT PRIMARY KEY,
                load_date TEXT NOT NULL,
                truck_no TEXT NOT NULL,
                trailer_no TEXT DEFAULT '',
                gas_source TEXT DEFAULT '',
                supplier TEXT DEFAULT '',
                station TEXT DEFAULT '',
                plan_arrive TEXT DEFAULT '',
                price REAL,
                net_weight REAL,
                amount REAL,
                driver_name TEXT DEFAULT '',
                driver_phone TEXT DEFAULT '',
                carrier TEXT DEFAULT '',
                note TEXT DEFAULT '',
                img_load TEXT DEFAULT '',
                img_unload TEXT DEFAULT '',
                img_waybill TEXT DEFAULT '',
                created_at TEXT,
                updated_at TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_plans_date ON plans(load_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_plans_truck ON plans(truck_no)")
        migrate_plan_arrive(conn)
        conn.commit()


def plan_to_dict(row):
    p = dict(row)

    def urlify(v):
        if not v:
            return ""
        parts = v.replace("\\", "/").split("/")
        if len(parts) >= 3 and parts[0] == "images":
            return "/images/" + "/".join(parts[1:])
        return "/images/" + os.path.basename(v)

    p["images"] = {
        "load": urlify(p.get("img_load")),
        "unload": urlify(p.get("img_unload")),
        "waybill": urlify(p.get("img_waybill")),
    }
    p["complete"] = all(bool(v) for v in p["images"].values())
    p["missing"] = [k for k, v in p["images"].items() if not v]
    for k in ("img_load", "img_unload", "img_waybill"):
        p[k] = urlify(p.get(k))
    return p


def now_str():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm_plate(s):
    if not s:
        return ""
    return re.sub(r"[^A-Z0-9\u4e00-\u9fff]", "", s.upper())


# ---------------- 文本解析 ----------------
def _today():
    return datetime.date.today()


def make_date(year, month, day):
    try:
        return datetime.date(int(year), int(month), int(day)).strftime("%Y-%m-%d")
    except ValueError:
        return ""


def parse_paste_text(text):
    """解析剪贴板粘贴的计划文本（微信等来源）。"""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    res = {
        "load_date": "",
        "truck_no": "",
        "trailer_no": "",
        "gas_source": "",
        "supplier": "",
        "station": "",
        "plan_arrive": "",
        "price": None,
        "driver_name": "",
        "driver_phone": "",
        "carrier": "",
        "note": "",
    }
    warnings = []
    arrive_day = None
    arrive_hour = None

    # 第一行：气源地-站点（可选 N号H点）；横杠后的内容即为站点名称。
    if lines:
        first = lines[0]
        route = re.match(r"^([^\s\-—－]+)\s*[\-—－]\s*(.+)$", first)
        if route:
            res["gas_source"] = route.group(1).strip()
            tail = route.group(2).strip()
            date_before_station = re.match(r"^(\d{1,2})[号日](?:(\d{1,2})[点时])?\s*(.+)$", tail)
            date_after_station = re.match(r"^(.*?)(\d{1,2})[号日](?:(\d{1,2})[点时])?$", tail)
            arrive = date_before_station or date_after_station
            res["station"] = (
                date_before_station.group(3) if date_before_station
                else date_after_station.group(1) if date_after_station else tail
            ).strip()
            if arrive:
                arrive_day = int(arrive.group(1) if date_before_station else arrive.group(2))
                arrive_hour = int(date_before_station.group(2) if date_before_station else arrive.group(3)) if (date_before_station.group(2) if date_before_station else arrive.group(3)) else None
        else:
            date_first = re.match(r"^([^\s\-—－]+)\s+(\d{1,2})[号日](?:(\d{1,2})[点时])?\s*(.+)$", first)
            if date_first:
                res["gas_source"] = date_first.group(1)
                res["station"] = date_first.group(4).strip()
                arrive_day = int(date_first.group(2))
                arrive_hour = int(date_first.group(3)) if date_first.group(3) else None
            else:
                m = re.match(r"^([^\s\-—－]+)\s+(.+?)(\d{1,2})[号日](?:(\d{1,2})[点时])?$", first)
                if m:
                    res["gas_source"] = m.group(1)
                    res["station"] = m.group(2)
                    arrive_day = int(m.group(3))
                    arrive_hour = int(m.group(4)) if m.group(4) else None

    for line in lines:
        if "供应商" in line:
            m = SUPPLIER_RE.search(line)
            if m:
                res["supplier"] = m.group(1)
            m = PRICE_RE.search(line)
            if m:
                res["price"] = float(m.group(1))
            continue
        if "装车" in line:
            m = CN_DAY_RE.search(line)
            if m:
                d = int(m.group(1))
                t = _today()
                res["load_date"] = make_date(t.year, t.month, d)
                warnings.append("装车日期未写月份，已按本月推断，请核对")
            continue
        if "价格" in line:
            m = PRICE_RE.search(line)
            if m:
                res["price"] = float(m.group(1))
        if arrive_day is None and "到站" in line:
            m = CN_DAY_RE.search(line)
            if m:
                arrive_day = int(m.group(1))
                mh = re.search(r"(\d{1,2})[点时]", line)
                arrive_hour = int(mh.group(1)) if mh else None

    # 计算计划到站日期（完整日期，月份按装车日期推断）
    if arrive_day is not None:
        base = res["load_date"] or _today().strftime("%Y-%m-%d")
        try:
            bd = datetime.date.fromisoformat(base)
        except ValueError:
            bd = _today()
        if arrive_day < bd.day:
            nm = bd.month + 1
            ny = bd.year + (1 if nm > 12 else 0)
            nm = 1 if nm > 12 else nm
        else:
            ny, nm = bd.year, bd.month
        date = make_date(ny, nm, arrive_day)
        if date:
            res["plan_arrive"] = (
                f"{date}T{arrive_hour:02d}:00" if arrive_hour else f"{date}T00:00"
            )
        warnings.append("计划到站日期已按装车日期推断月份，请核对")

    # 车号/挂车号
    plates = PLATE_RE.findall(text)
    heads = [p for p in plates if not p.endswith("挂")]
    trailers = [p for p in plates if p.endswith("挂")]
    if heads:
        res["truck_no"] = heads[0]
    if trailers:
        res["trailer_no"] = trailers[0]

    # 司机：姓名/身份证/电话 成组
    drivers = []
    for i, ln in enumerate(lines):
        if ID_RE.fullmatch(ln):
            name = ""
            phone = ""
            if i > 0 and NAME_RE.fullmatch(lines[i - 1]):
                name = lines[i - 1]
            if i + 1 < len(lines) and PHONE_RE.fullmatch(lines[i + 1]):
                phone = lines[i + 1]
            drivers.append((name, phone))
    if drivers:
        res["driver_name"], res["driver_phone"] = drivers[0]
        if len(drivers) > 1:
            extra = "；".join(
                f"押运：{n} {ph}".strip() for n, ph in drivers[1:] if n or ph
            )
            if extra:
                res["note"] = extra

    # 优先识别“驾驶员：姓名 / 电话：号码”这类带标签格式。
    driver_index = next((i for i, ln in enumerate(lines) if re.match(r"^(?:驾驶员|司机)(?:姓名)?\s*[：:]", ln)), -1)
    if driver_index >= 0:
        name = re.sub(r"^(?:驾驶员|司机)(?:姓名)?\s*[：:]\s*", "", lines[driver_index]).strip()
        if name:
            res["driver_name"] = name
        for ln in lines[driver_index + 1:]:
            if re.match(r"^(?:押运员|押运|供应商)\s*[：:]", ln):
                break
            phone = re.fullmatch(r"(?:电话(?:号码)?\s*[：:]?\s*)?(1\d{10})\s*", ln)
            if phone:
                res["driver_phone"] = phone.group(1)
                break

    # 承运公司
    m = COMPANY_RE.search(text)
    if m:
        res["carrier"] = m.group(1)

    # 常用站点在任意粘贴行中出现时，优先作为站点字段。
    for line in lines:
        station = next((name for name in COMMON_STATIONS if name in line), None)
        if station:
            res["station"] = station
            break

    return {"fields": res, "warnings": warnings}


def migrate_plan_arrive(conn):
    """把旧文本型/日期型计划到站日期迁移为 YYYY-MM-DDTHH:MM。"""
    rows = conn.execute(
        "SELECT id, load_date, plan_arrive, note FROM plans"
    ).fetchall()
    for r in rows:
        pa = r["plan_arrive"] or ""
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}", pa):
            continue
        date = normalize_arrive(pa, r["load_date"])
        if not date:
            continue
        # 旧数据把时刻写在备注里（如"计划到站19点"），迁移时恢复到时间
        if not re.search(r"[T ]\d{1,2}:\d{2}", pa):
            mn = re.search(r"计划到站(\d{1,2})点", r["note"] or "")
            if mn:
                date = f"{date[:11]}{int(mn.group(1)):02d}:00"
        conn.execute(
            "UPDATE plans SET plan_arrive=? WHERE id=?",
            (date, r["id"]),
        )


def normalize_arrive(plan_arrive, load_date):
    """把计划到站日期统一为 datetime-local 格式 YYYY-MM-DDTHH:MM；无法识别返回空。"""
    pa = (plan_arrive or "").strip()
    if not pa:
        return ""
    m = re.match(r"^(\d{4}-\d{2}-\d{2})[T ](\d{1,2})(?::(\d{2}))?$", pa)
    if m:
        hour = int(m.group(2))
        minute = int(m.group(3) or 0)
        return f"{m.group(1)}T{hour:02d}:{minute:02d}"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", pa):
        return f"{pa}T00:00"
    hour = None
    date = ""
    m = FULL_DATE_RE.search(pa)
    if m:
        date = make_date(m.group(1), m.group(2), m.group(3))
        mh = re.search(r"(\d{1,2})[点时]", pa)
        hour = int(mh.group(1)) if mh else None
    else:
        m = CN_DAY_RE.search(pa)
        if m:
            day = int(m.group(1))
            mh = re.search(r"(\d{1,2})[点时]", pa)
            hour = int(mh.group(1)) if mh else None
            try:
                bd = datetime.date.fromisoformat(load_date or "")
            except ValueError:
                bd = _today()
            if day < bd.day:
                nm = bd.month + 1
                ny = bd.year + (1 if nm > 12 else 0)
                nm = 1 if nm > 12 else nm
            else:
                ny, nm = bd.year, bd.month
            date = make_date(ny, nm, day)
    if not date:
        return ""
    return f"{date}T{hour:02d}:00" if hour else f"{date}T00:00"


def parse_ocr_text(text):
    """从单据 OCR 文本中提取车号、日期、单据类型、净重。"""
    parsed = {
        "truck_no": "",
        "load_date": "",
        "doc_type": "",
        "doc_type_confidence": "low",
        "net_weight": None,
    }
    plates = PLATE_RE.findall(text)
    heads = [p for p in plates if not p.endswith("挂")]
    if heads:
        parsed["truck_no"] = norm_plate(heads[0])

    m = FULL_DATE_RE.search(text)
    if m:
        parsed["load_date"] = make_date(m.group(1), m.group(2), m.group(3))
    else:
        m = CN_DAY_RE.search(text)
        if m:
            t = _today()
            parsed["load_date"] = make_date(t.year, t.month, int(m.group(1)))

    # 单据类型：先看显式标题，再按关键词打分
    low = text.lower()
    scores = {"load": 0, "unload": 0, "waybill": 0}
    if "装车磅单" in text or "装车计量单" in text:
        scores["load"] += 6
    if "卸车磅单" in text or "卸车计量单" in text:
        scores["unload"] += 6
    if "运单" in text and "磅单" not in text:
        scores["waybill"] += 6
    kw = {
        "load": ["装车", "装货", "发货单位", "发货方", "出厂", "发运", "皮重"],
        "unload": ["卸车", "卸货", "收货单位", "收货方", "回皮", "实收", "结算"],
        "waybill": ["托运", "承运", "运输合同", "调度", "起运地", "到达地", "货主"],
    }
    for k, words in kw.items():
        for w in words:
            if w in text:
                scores[k] += 1
    best = max(scores, key=lambda k: scores[k])
    if scores[best] >= 2:
        parsed["doc_type"] = best
        parsed["doc_type_confidence"] = "high" if scores[best] >= 4 else "medium"

    m = NET_WEIGHT_RE.search(text)
    if m:
        val = float(m.group(1))
        unit = (m.group(2) or "").lower()
        if unit in ("kg", "千克"):
            val = val / 1000.0
        parsed["net_weight"] = round(val, 3)
    return parsed


def find_candidates(truck_no, load_date, limit=8):
    if not truck_no:
        return []
    conn = get_db()
    try:
        if load_date:
            rows = conn.execute(
                "SELECT * FROM plans WHERE truck_no=? AND load_date=? ORDER BY load_date DESC LIMIT ?",
                (truck_no, load_date, limit),
            ).fetchall()
            if not rows:
                rows = conn.execute(
                    "SELECT * FROM plans WHERE truck_no=? ORDER BY load_date DESC LIMIT ?",
                    (truck_no, limit),
                ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM plans WHERE truck_no=? ORDER BY load_date DESC LIMIT ?",
                (truck_no, limit),
            ).fetchall()
        return [plan_to_dict(r) for r in rows]
    finally:
        conn.close()


# ---------------- 图片处理 ----------------
def save_image_bytes(plan_id, doc_type, raw_bytes):
    """保存图片为 JPEG，返回相对路径（相对 data/）。"""
    plan_dir = os.path.join(IMAGES_DIR, plan_id)
    os.makedirs(plan_dir, exist_ok=True)
    tmp_path = os.path.join(plan_dir, f"{doc_type}_tmp_{uuid.uuid4().hex}")
    with open(tmp_path, "wb") as f:
        f.write(raw_bytes)
    final_path = os.path.join(plan_dir, f"{doc_type}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.jpg")
    r = subprocess.run(
        ["sips", "-s", "format", "jpeg", tmp_path, "--out", final_path],
        capture_output=True,
    )
    if r.returncode != 0 or not os.path.exists(final_path):
        # sips 失败则原样保存
        shutil.move(tmp_path, final_path)
    else:
        os.remove(tmp_path)
    return os.path.relpath(final_path, DATA_DIR)


def remove_plan_images(plan_id):
    d = os.path.join(IMAGES_DIR, plan_id)
    if os.path.isdir(d):
        shutil.rmtree(d, ignore_errors=True)


# ---------------- OCR ----------------
def ensure_ocr():
    if os.path.exists(OCR_BIN):
        return True
    if not os.path.exists(OCR_SRC):
        return False
    try:
        r = subprocess.run(
            [
                "clang",
                "-fobjc-arc",
                "-framework",
                "Foundation",
                "-framework",
                "Vision",
                "-framework",
                "AppKit",
                "-o",
                OCR_BIN,
                OCR_SRC,
            ],
            capture_output=True,
            text=True,
            timeout=180,
        )
        return r.returncode == 0
    except Exception:
        return False


def run_ocr(image_path):
    import time

    last_err = ""
    for attempt in range(3):
        try:
            r = subprocess.run(
                [OCR_BIN, image_path], capture_output=True, text=True, timeout=60
            )
            if r.returncode == 0 and r.stdout.strip():
                return r.stdout.strip(), ""
            last_err = r.stderr.strip() or "OCR 未识别出文字"
        except Exception as e:
            last_err = str(e)
        if attempt < 2:
            time.sleep(0.5)
    return "", last_err


# ---------------- Excel 导出 ----------------
def export_xlsx(rows, from_date, to_date):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    headers = [
        "气源地", "装车日期", "车号", "供应商", "站点", "计划到站日期",
        "价格(元/吨)", "装车磅单", "卸车磅单", "运单",
        "净重(吨)", "金额(元)", "备注",
    ]
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "对账清单"

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    row = 2
    total_amount = 0.0
    total_weight = 0.0
    for p in rows:
        price = p.get("price") or 0
        weight = p.get("net_weight") or 0
        amount = round(price * weight, 2) if price and weight else None
        total_amount += amount or 0
        total_weight += weight or 0
        values = [
            p.get("gas_source") or "",
            p.get("load_date") or "",
            p.get("truck_no") or "",
            p.get("supplier") or "",
            p.get("station") or "",
            p.get("plan_arrive") or "",
            price if price else None,
            None, None, None,
            weight if weight else None,
            amount,
            p.get("note") or "",
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
        ld = p.get("load_date") or ""
        try:
            d = datetime.date.fromisoformat(ld)
            ws.cell(row=row, column=2, value=d)
        except ValueError:
            pass
        pa = p.get("plan_arrive") or ""
        try:
            d = datetime.datetime.fromisoformat(pa.replace(" ", "T"))
            ws.cell(row=row, column=6, value=d)
        except ValueError:
            pass
        ws.cell(row=row, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=6).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=row, column=7).number_format = "#,##0"
        ws.cell(row=row, column=11).number_format = "#,##0.00"
        ws.cell(row=row, column=12).number_format = "#,##0.00"
        ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=9).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=10).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 70

        for c, key in ((8, "img_load"), (9, "img_unload"), (10, "img_waybill")):
            url = p.get(key) or ""
            if url:
                path = os.path.join(
                    IMAGES_DIR, p.get("id") or "", os.path.basename(url)
                )
                if os.path.exists(path):
                    try:
                        img = XLImage(path)
                        img.width = 90
                        img.height = 62
                        img.anchor = f"{get_column_letter(c)}{row}"
                        ws.add_image(img)
                    except Exception:
                        ws.cell(row=row, column=c, value="图")
        row += 1

    # 合计行
    if rows:
        cell = ws.cell(row=row, column=3, value=f"合计 {len(rows)} 车")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=11, value=round(total_weight, 2)).font = Font(bold=True)
        ws.cell(row=row, column=11).number_format = "#,##0.00"
        ws.cell(row=row, column=12, value=round(total_amount, 2)).font = Font(bold=True)
        ws.cell(row=row, column=12).number_format = "#,##0.00"

    widths = [12, 12, 12, 12, 14, 12, 11, 13, 13, 13, 10, 12, 28]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # 汇总 sheet：按天统计车数
    ws2 = wb.create_sheet("每日汇总")
    h2 = ["日期", "车数", "单据齐全", "缺装车磅单", "缺卸车磅单", "缺运单", "金额合计(元)"]
    for c, h in enumerate(h2, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    day_map = {}
    for p in rows:
        d = (p.get("load_date") or "")[:10]
        day_map.setdefault(d, []).append(p)
    r2 = 2
    grand_total = 0
    for d in sorted(day_map):
        ps = day_map[d]
        miss_load = sum(1 for x in ps if not x.get("img_load"))
        miss_unload = sum(1 for x in ps if not x.get("img_unload"))
        miss_waybill = sum(1 for x in ps if not x.get("img_waybill"))
        complete = sum(1 for x in ps if x.get("complete"))
        amount = sum(
            round((x.get("price") or 0) * (x.get("net_weight") or 0), 2) for x in ps
        )
        grand_total += len(ps)
        vals = [d, len(ps), complete, miss_load, miss_unload, miss_waybill, round(amount, 2)]
        for c, v in enumerate(vals, start=1):
            cell = ws2.cell(row=r2, column=c, value=v)
            cell.border = border
        ws2.cell(row=r2, column=1).number_format = "yyyy-mm-dd"
        ws2.cell(row=r2, column=7).number_format = "#,##0.00"
        r2 += 1
    cell = ws2.cell(row=r2, column=2, value=grand_total)
    cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 12
    for c in "BCDEFG":
        ws2.column_dimensions[c].width = 13
    ws2.freeze_panes = "A2"

    title = f"对账清单_{from_date}_{to_date}.xlsx"
    buf = __import__("io").BytesIO()
    wb.save(buf)
    buf.seek(0)
    return title, buf


# ---------------- 查询 ----------------
def query_plans(from_date, to_date, q, date_field="load_date", supplier=""):
    """查询计划；date_field 仅允许按装车日期或计划到站日期筛选。"""
    conn = get_db()
    try:
        sql = "SELECT * FROM plans WHERE 1=1"
        args = []
        date_expr = "substr(plan_arrive, 1, 10)" if date_field == "plan_arrive" else "load_date"
        if from_date:
            sql += f" AND {date_expr} >= ?"
            args.append(from_date)
        if to_date:
            sql += f" AND {date_expr} <= ?"
            args.append(to_date)
        if supplier:
            sql += " AND supplier LIKE ?"
            args.append(f"%{supplier}%")
        if q:
            sql += """ AND (truck_no LIKE ? OR gas_source LIKE ? OR supplier LIKE ?
                         OR station LIKE ? OR driver_name LIKE ? OR note LIKE ?)"""
            like = f"%{q}%"
            args += [like] * 6
        sql += f" ORDER BY {date_expr} DESC, created_at DESC"
        rows = conn.execute(sql, args).fetchall()
        return [plan_to_dict(r) for r in rows]
    finally:
        conn.close()


def get_plan(plan_id):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM plans WHERE id=?", (plan_id,)).fetchone()
        return plan_to_dict(row) if row else None
    finally:
        conn.close()


def compute_amount(price, weight):
    try:
        if price and weight:
            return round(float(price) * float(weight), 2)
    except (TypeError, ValueError):
        pass
    return None


def upsert_plan(data, plan_id=None):
    """新增或更新计划。data 中可含 img_*_b64（base64 字符串，'' 表示删除图片）。"""
    load_date = (data.get("load_date") or "").strip()
    truck_no = norm_plate(data.get("truck_no") or "").strip() or (data.get("truck_no") or "").strip()
    if not load_date or not truck_no:
        raise ValueError("装车日期和车号必填")

    def num(k):
        v = data.get(k)
        if v in (None, ""):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    price = num("price")
    weight = num("net_weight")
    amount = compute_amount(price, weight)
    if "amount" in data and data.get("amount") not in (None, ""):
        try:
            amount = float(data["amount"])
        except (TypeError, ValueError):
            pass

    arrive = normalize_arrive(data.get("plan_arrive"), load_date)
    if not arrive:
        raise ValueError("计划到站日期必填")

    fields = {
        "load_date": load_date,
        "truck_no": truck_no,
        "trailer_no": (data.get("trailer_no") or "").strip(),
        "gas_source": (data.get("gas_source") or "").strip(),
        "supplier": (data.get("supplier") or "").strip(),
        "station": (data.get("station") or "").strip(),
        "plan_arrive": arrive,
        "price": price,
        "net_weight": weight,
        "amount": amount,
        "driver_name": (data.get("driver_name") or "").strip(),
        "driver_phone": (data.get("driver_phone") or "").strip(),
        "carrier": (data.get("carrier") or "").strip(),
        "note": (data.get("note") or "").strip(),
    }
    conn = get_db()
    try:
        if plan_id:
            existing = conn.execute("SELECT * FROM plans WHERE id=?", (plan_id,)).fetchone()
            if not existing:
                raise ValueError("计划不存在")
            fields["updated_at"] = now_str()
            sets = ", ".join(f"{k}=?" for k in fields)
            conn.execute(
                f"UPDATE plans SET {sets} WHERE id=?",
                list(fields.values()) + [plan_id],
            )
        else:
            plan_id = uuid.uuid4().hex
            fields["id"] = plan_id
            fields["created_at"] = now_str()
            fields["updated_at"] = now_str()
            cols = ", ".join(fields.keys())
            marks = ", ".join("?" for _ in fields)
            conn.execute(
                f"INSERT INTO plans ({cols}) VALUES ({marks})", list(fields.values())
            )
        # 图片处理
        for doc_type in DOC_TYPES:
            key = f"img_{doc_type}_b64"
            if key in data:
                col = IMG_COLUMNS[doc_type]
                if data[key] == "":
                    old = conn.execute(
                        f"SELECT {col} FROM plans WHERE id=?", (plan_id,)
                    ).fetchone()
                    if old and old[col]:
                        op = os.path.join(DATA_DIR, old[col])
                        if os.path.exists(op):
                            os.remove(op)
                    conn.execute(f"UPDATE plans SET {col}='' WHERE id=?", (plan_id,))
                else:
                    try:
                        raw = base64.b64decode(data[key].split(",")[-1])
                        rel = save_image_bytes(plan_id, doc_type, raw)
                        conn.execute(
                            f"UPDATE plans SET {col}=? WHERE id=?", (rel, plan_id)
                        )
                    except Exception as e:
                        raise ValueError(f"图片保存失败（{doc_type}）：{e}")
        conn.commit()
        return get_plan(plan_id)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def delete_plan(plan_id):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM plans WHERE id=?", (plan_id,)).fetchone()
        if not row:
            return False
        conn.execute("DELETE FROM plans WHERE id=?", (plan_id,))
        conn.commit()
        remove_plan_images(plan_id)
        return True
    finally:
        conn.close()


def stats_day(day):
    rows = query_plans(day, day, "", "plan_arrive")
    return {
        "date": day,
        "total": len(rows),
        "complete": sum(1 for p in rows if p["complete"]),
        "missing_load": sum(1 for p in rows if not p["images"]["load"]),
        "missing_unload": sum(1 for p in rows if not p["images"]["unload"]),
        "missing_waybill": sum(1 for p in rows if not p["images"]["waybill"]),
        "amount": round(
            sum(
                (p.get("price") or 0) * (p.get("net_weight") or 0) for p in rows
            ),
            2,
        ),
        "plans": rows,
    }


def stats_month(month):
    import calendar

    from_date = f"{month}-01"
    try:
        last = calendar.monthrange(int(month[:4]), int(month[5:7]))[1]
        to_date = f"{month}-{last:02d}"
    except (ValueError, IndexError):
        to_date = f"{month}-31"
    rows = query_plans(from_date, to_date, "", "plan_arrive")
    days = {}
    for p in rows:
        days.setdefault((p.get("plan_arrive") or "")[:10], []).append(p)
    day_stats = []
    for d in sorted(days):
        ps = days[d]
        day_stats.append(
            {
                "date": d,
                "total": len(ps),
                "complete": sum(1 for p in ps if p["complete"]),
                "missing_load": sum(1 for p in ps if not p["images"]["load"]),
                "missing_unload": sum(1 for p in ps if not p["images"]["unload"]),
                "missing_waybill": sum(1 for p in ps if not p["images"]["waybill"]),
                "amount": round(
                    sum((p.get("price") or 0) * (p.get("net_weight") or 0) for p in ps),
                    2,
                ),
            }
        )
    return {
        "month": month,
        "total": len(rows),
        "complete": sum(1 for p in rows if p["complete"]),
        "missing_load": sum(1 for p in rows if not p["images"]["load"]),
        "missing_unload": sum(1 for p in rows if not p["images"]["unload"]),
        "missing_waybill": sum(1 for p in rows if not p["images"]["waybill"]),
        "amount": round(
            sum((p.get("price") or 0) * (p.get("net_weight") or 0) for p in rows),
            2,
        ),
        "days": day_stats,
    }


def stats_range(from_date, to_date, supplier=""):
    rows = query_plans(from_date, to_date, "", "plan_arrive", supplier)
    days = {}
    for p in rows:
        days.setdefault((p.get("plan_arrive") or "")[:10], []).append(p)
    day_stats = []
    for d in sorted(days):
        ps = days[d]
        day_stats.append(
            {
                "date": d,
                "total": len(ps),
                "complete": sum(1 for p in ps if p["complete"]),
                "missing_load": sum(1 for p in ps if not p["images"]["load"]),
                "missing_unload": sum(1 for p in ps if not p["images"]["unload"]),
                "missing_waybill": sum(1 for p in ps if not p["images"]["waybill"]),
                "amount": round(
                    sum((p.get("price") or 0) * (p.get("net_weight") or 0) for p in ps),
                    2,
                ),
                "plans": ps,
            }
        )
    return {
        "from": from_date,
        "to": to_date,
        "total": len(rows),
        "complete": sum(1 for p in rows if p["complete"]),
        "missing_load": sum(1 for p in rows if not p["images"]["load"]),
        "missing_unload": sum(1 for p in rows if not p["images"]["unload"]),
        "missing_waybill": sum(1 for p in rows if not p["images"]["waybill"]),
        "amount": round(
            sum((p.get("price") or 0) * (p.get("net_weight") or 0) for p in rows),
            2,
        ),
        "days": day_stats,
    }


# ---------------- HTTP 服务 ----------------
class Handler(BaseHTTPRequestHandler):
    server_version = "PlanTool/1.0"

    def log_message(self, fmt, *args):
        sys.stderr.write("%s - %s\n" % (self.address_string(), fmt % args))

    def _send(self, code, body=b"", ctype="application/octet-stream", headers=None):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        if headers:
            for k, v in headers.items():
                self.send_header(k, v)
        self.end_headers()
        if body:
            self.wfile.write(body)

    def send_json(self, obj, code=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self._send(code, body, "application/json; charset=utf-8")

    def send_error_json(self, msg, code=400):
        self.send_json({"ok": False, "error": msg}, code)

    def read_json(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        try:
            return json.loads(raw.decode("utf-8"))
        except Exception:
            return {}

    def _serve_file(self, path, root):
        full = os.path.realpath(os.path.join(root, path))
        root_real = os.path.realpath(root)
        if not full.startswith(root_real + os.sep) and full != root_real:
            self.send_error_json("路径不允许", 403)
            return
        if not os.path.isfile(full):
            self.send_error(404)
            return
        ctype = mimetypes.guess_type(full)[0] or "application/octet-stream"
        with open(full, "rb") as f:
            self._send(200, f.read(), ctype)

    # ----- GET -----
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        qs = parse_qs(parsed.query)

        if path == "/" or path == "/index.html":
            self._serve_file("index.html", STATIC_DIR)
            return
        if path.startswith("/static/"):
            self._serve_file(path[len("/static/"):], STATIC_DIR)
            return
        if path.startswith("/images/"):
            self._serve_file(path[len("/images/"):], IMAGES_DIR)
            return
        if path == "/api/health":
            self.send_json({"ok": True, "ocr": os.path.exists(OCR_BIN)})
            return
        if path == "/api/state":
            today = _today().strftime("%Y-%m-%d")
            month = _today().strftime("%Y-%m")
            self.send_json(
                {
                    "ok": True,
                    "today": stats_day(today),
                    "month": stats_month(month),
                    "ocr_available": os.path.exists(OCR_BIN),
                }
            )
            return
        if path == "/api/plans":
            from_date = (qs.get("from") or [""])[0]
            to_date = (qs.get("to") or [""])[0]
            q = (qs.get("q") or [""])[0]
            date_field = (qs.get("date_field") or ["load_date"])[0]
            self.send_json({"ok": True, "plans": query_plans(from_date, to_date, q, date_field)})
            return
        if path.startswith("/api/plans/"):
            plan_id = path[len("/api/plans/"):].strip("/")
            p = get_plan(plan_id)
            if not p:
                self.send_error_json("计划不存在", 404)
            else:
                self.send_json({"ok": True, "plan": p})
            return
        if path == "/api/stats":
            mode = (qs.get("mode") or ["day"])[0]
            if mode == "month":
                month = (qs.get("month") or [_today().strftime("%Y-%m")])[0]
                self.send_json({"ok": True, "stats": stats_month(month)})
            elif mode == "range":
                from_date = (qs.get("from") or [""])[0]
                to_date = (qs.get("to") or [""])[0]
                supplier = (qs.get("supplier") or [""])[0]
                self.send_json({"ok": True, "stats": stats_range(from_date, to_date, supplier)})
            else:
                day = (qs.get("date") or [_today().strftime("%Y-%m-%d")])[0]
                self.send_json({"ok": True, "stats": stats_day(day)})
            return
        if path == "/api/export":
            from_date = (qs.get("from") or [""])[0]
            to_date = (qs.get("to") or [""])[0]
            supplier = (qs.get("supplier") or [""])[0]
            rows = query_plans(from_date, to_date, "", "plan_arrive", supplier)
            title, buf = export_xlsx(rows, from_date or "全部", to_date or "全部")
            encoded = urllib_quote(title)
            self._send(
                200,
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                {"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
            )
            return
        self.send_error_json("未找到接口", 404)

    # ----- POST -----
    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        data = self.read_json()

        if path == "/api/plans":
            try:
                plan = upsert_plan(data)
            except ValueError as e:
                self.send_error_json(str(e))
                return
            self.send_json({"ok": True, "plan": plan}, 201)
            return
        if path.startswith("/api/plans/"):
            plan_id = path[len("/api/plans/"):].strip("/")
            try:
                plan = upsert_plan(data, plan_id)
            except ValueError as e:
                self.send_error_json(str(e))
                return
            self.send_json({"ok": True, "plan": plan})
            return
        if path == "/api/parse_text":
            text = data.get("text") or ""
            result = parse_paste_text(text)
            self.send_json({"ok": True, **result})
            return
        if path == "/api/ocr":
            b64 = data.get("image_base64") or ""
            if not b64:
                self.send_error_json("缺少图片")
                return
            try:
                raw = base64.b64decode(b64.split(",")[-1])
            except Exception:
                self.send_error_json("图片数据无效")
                return
            tmp = os.path.join(IMAGES_DIR, f"ocr_tmp_{uuid.uuid4().hex}")
            with open(tmp, "wb") as f:
                f.write(raw)
            try:
                text, err = run_ocr(tmp)
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)
            if not text:
                self.send_json(
                    {
                        "ok": False,
                        "error": err or "OCR 未识别出文字",
                        "text": "",
                        "parsed": {},
                        "candidates": [],
                    }
                )
                return
            parsed = parse_ocr_text(text)
            candidates = find_candidates(parsed["truck_no"], parsed["load_date"])
            self.send_json(
                {
                    "ok": True,
                    "text": text,
                    "parsed": parsed,
                    "candidates": candidates,
                }
            )
            return
        if path == "/api/attach":
            plan_id = data.get("plan_id") or ""
            doc_type = data.get("doc_type") or ""
            b64 = data.get("image_base64") or ""
            if doc_type not in DOC_TYPES:
                self.send_error_json("单据类型无效")
                return
            if not plan_id or not b64:
                self.send_error_json("缺少计划或图片")
                return
            try:
                raw = base64.b64decode(b64.split(",")[-1])
                rel = save_image_bytes(plan_id, doc_type, raw)
            except Exception as e:
                self.send_error_json(f"图片保存失败：{e}")
                return
            conn = get_db()
            try:
                col = IMG_COLUMNS[doc_type]
                old = conn.execute(
                    f"SELECT {col} FROM plans WHERE id=?", (plan_id,)
                ).fetchone()
                if old and old[col]:
                    op = os.path.join(DATA_DIR, old[col])
                    if os.path.exists(op):
                        os.remove(op)
                conn.execute(
                    f"UPDATE plans SET {col}=?, updated_at=? WHERE id=?",
                    (rel, now_str(), plan_id),
                )
                # 卸车磅单识别出净重时自动补全
                weight = data.get("net_weight")
                if doc_type == "unload" and weight:
                    try:
                        weight = float(weight)
                    except (TypeError, ValueError):
                        weight = None
                    if weight:
                        row = conn.execute(
                            "SELECT price FROM plans WHERE id=?", (plan_id,)
                        ).fetchone()
                        price = row["price"] if row else None
                        amount = compute_amount(price, weight)
                        conn.execute(
                            "UPDATE plans SET net_weight=?, amount=?, updated_at=? WHERE id=?",
                            (weight, amount, now_str(), plan_id),
                        )
                conn.commit()
            finally:
                conn.close()
            self.send_json({"ok": True, "plan": get_plan(plan_id)})
            return
        self.send_error_json("未找到接口", 404)

    def do_PUT(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/plans/"):
            plan_id = parsed.path[len("/api/plans/"):].strip("/")
            data = self.read_json()
            try:
                plan = upsert_plan(data, plan_id)
            except ValueError as e:
                self.send_error_json(str(e))
                return
            self.send_json({"ok": True, "plan": plan})
            return
        self.send_error_json("未找到接口", 404)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/plans/"):
            plan_id = parsed.path[len("/api/plans/"):].strip("/")
            if delete_plan(plan_id):
                self.send_json({"ok": True})
            else:
                self.send_error_json("计划不存在", 404)
            return
        self.send_error_json("未找到接口", 404)

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()


def urllib_quote(s):
    from urllib.parse import quote
    return quote(s)


def main():
    init_db()
    ocr_ok = ensure_ocr()
    if not ocr_ok:
        print("警告：OCR 编译失败，图片自动识别将不可用", file=sys.stderr)
    else:
        print("OCR 就绪")
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"车辆计划对账系统已启动：http://127.0.0.1:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
